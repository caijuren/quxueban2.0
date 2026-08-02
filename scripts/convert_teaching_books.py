#!/usr/bin/env python3
"""把 上海小学教辅书目主表.xlsx 转换成 JSON，供 Prisma seed 脚本导入。"""

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
EXCEL_FILE = ROOT / "上海小学教辅书目主表.xlsx"
OUTPUT_DIR = ROOT / "prisma" / "seed-data"

GRADE_ORDER = [
    "一年级上", "一年级下",
    "二年级上", "二年级下",
    "三年级上", "三年级下",
    "四年级上", "四年级下",
    "五年级上", "五年级下",
    "跨年级通用",
]


def normalize_grade(raw: str | None) -> str:
    if not raw:
        return "跨年级通用"
    raw = str(raw).strip().replace(" ", "")
    # 常见变体统一
    mapping = {
        "一年级第一学期": "一年级上",
        "一年级第二学期": "一年级下",
        "二年级第一学期": "二年级上",
        "二年级第二学期": "二年级下",
        "三年级第一学期": "三年级上",
        "三年级第二学期": "三年级下",
        "四年级第一学期": "四年级上",
        "四年级第二学期": "四年级下",
        "五年级第一学期": "五年级上",
        "五年级第二学期": "五年级下",
    }
    return mapping.get(raw, raw)


def normalize_new_textbook(raw: str | None) -> str:
    if not raw:
        return "否"
    raw = str(raw).strip()
    if raw in ("是", "否", "部分适配"):
        return raw
    if "部分" in raw:
        return "部分适配"
    if raw in ("Y", "y", "Yes", "yes", "true", "True", "是"):
        return "是"
    return "否"


def validate_isbn(raw: str | None) -> str | None:
    if not raw:
        return None
    s = str(raw).strip().replace("-", "").replace(" ", "")
    if re.fullmatch(r"978\d{10}", s):
        return s
    return None


def parse_price(raw) -> float | None:
    if raw is None:
        return None
    try:
        return float(raw)
    except (ValueError, TypeError):
        return None


def load_publishers(wb) -> list[dict]:
    ws = wb["出版社对照表"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data = []
    for row in rows[1:]:
        if not row or not row[1]:
            continue
        item = dict(zip(headers, row))
        data.append({
            "name": item.get("出版社名称", "").strip(),
            "shortName": (item.get("简称") or "").strip() or None,
            "website": (item.get("网址") or "").strip() or None,
            "strongSubjects": (item.get("上海小学强势学科") or "").strip() or None,
            "series": (item.get("代表教辅系列") or "").strip() or None,
            "contactEmail": (item.get("图片授权/合作邮箱") or "").strip() or None,
            "note": (item.get("备注") or "").strip() or None,
        })
    return data


def load_content_types(wb) -> list[dict]:
    ws = wb["内容类型分类表"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data = []
    for row in rows[1:]:
        if not row or not row[1]:
            continue
        item = dict(zip(headers, row))
        data.append({
            "name": item.get("内容类型", "").strip(),
            "description": (item.get("定义") or "").strip() or None,
            "typicalExamples": (item.get("典型代表（上海小学）") or "").strip() or None,
            "usageScenario": (item.get("使用场景") or "").strip() or None,
            "difficultyRange": (item.get("难度区间") or "").strip() or None,
        })
    return data


def load_books(wb) -> list[dict]:
    ws = wb["书目主表"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data = []
    for idx, row in enumerate(rows[1:], start=2):
        if not row or not row[0]:
            continue
        item = dict(zip(headers, row))

        book_id = str(item.get("书目ID") or "").strip()
        title = str(item.get("书名（完整）") or "").strip()
        publisher = str(item.get("出版社") or "").strip()
        subject = str(item.get("学科") or "").strip()
        content_type = str(item.get("内容类型") or "").strip()
        grade = normalize_grade(item.get("适用年级"))
        status = str(item.get("上架状态") or "在售").strip()

        if not book_id or not title or not publisher or not subject or not content_type:
            print(f"跳过第 {idx} 行：关键字段缺失")
            continue

        # 只导入在售和待核验的，跳过完全空行
        if status not in ("在售", "待核验"):
            continue

        verified_at = item.get("最后核验日期")
        if isinstance(verified_at, str):
            verified_at = verified_at.strip() or None

        isbn = validate_isbn(item.get("ISBN"))

        book = {
            "bookId": book_id,
            "title": title,
            "isbn": isbn,
            "publisher": publisher,
            "author": (str(item.get("作者/编写组") or "").strip()) or None,
            "editionDate": (str(item.get("出版日期") or "").strip()) or None,
            "editionNumber": (str(item.get("版次") or "").strip()) or None,
            "price": parse_price(item.get("定价（元）")),
            "subject": subject,
            "contentType": content_type,
            "grade": grade,
            "textbookVersion": (str(item.get("适配教材版本") or "").strip()) or None,
            "isNewTextbook": normalize_new_textbook(item.get("是否新教材适配")),
            "difficulty": int(item.get("难度星级") or 1),
            "targetAudience": (str(item.get("适合人群") or "").strip()) or None,
            "sellingPoints": (str(item.get("核心卖点") or "").strip()) or None,
            "structureDesc": (str(item.get("结构说明") or "").strip()) or None,
            "companionSuggestion": (str(item.get("配套建议") or "").strip()) or None,
            "coverImageUrl": (str(item.get("封面图URL") or "").strip()) or None,
            "jdUrl": (str(item.get("京东购买链接") or "").strip()) or None,
            "dangdangUrl": (str(item.get("当当购买链接") or "").strip()) or None,
            "officialUrl": (str(item.get("官方旗舰店链接") or "").strip()) or None,
            "status": status,
            "lastVerifiedAt": verified_at,
            "note": (str(item.get("备注") or "").strip()) or None,
        }
        data.append(book)

    # 按年级、学科、书名排序
    data.sort(key=lambda b: (
        GRADE_ORDER.index(b["grade"]) if b["grade"] in GRADE_ORDER else 99,
        b["subject"],
        b["title"],
    ))
    return data


def merge_missing_publishers(publishers: list[dict], books: list[dict]) -> list[dict]:
    """把书目里出现但对照表里没有的出版社补齐，避免导入时跳过。"""
    existing_names = {p["name"] for p in publishers}
    missing_names = sorted({b["publisher"] for b in books if b["publisher"] not in existing_names})
    for name in missing_names:
        publishers.append({
            "name": name,
            "shortName": None,
            "website": None,
            "strongSubjects": None,
            "series": None,
            "contactEmail": None,
            "note": "从书目主表自动补齐",
        })
    return publishers


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)

    publishers = load_publishers(wb)
    content_types = load_content_types(wb)
    books = load_books(wb)

    publishers = merge_missing_publishers(publishers, books)

    (OUTPUT_DIR / "publishers.json").write_text(
        json.dumps(publishers, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (OUTPUT_DIR / "contentTypes.json").write_text(
        json.dumps(content_types, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (OUTPUT_DIR / "books.json").write_text(
        json.dumps(books, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print(f"出版社：{len(publishers)} 条")
    print(f"内容类型：{len(content_types)} 条")
    print(f"书目：{len(books)} 条")
    print(f"输出目录：{OUTPUT_DIR}")


if __name__ == "__main__":
    main()
